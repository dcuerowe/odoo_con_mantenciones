import io
import openpyxl
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter
from config import EXCEL_URL

def modify_excel_file(resumen, sheet_name, table_name, sharepoint_client):
    # Descarga el archivo Excel desde SharePoint
    excel = sharepoint_client.download_file(EXCEL_URL)
    if excel:
        try:
            # Convierte los bytes descargados en un objeto BytesIO para manipulación en memoria
            excel_file = io.BytesIO(excel)
            # Carga el archivo Excel en openpyxl
            wl = openpyxl.load_workbook(excel_file)
            # Selecciona la hoja de trabajo especificada
            wh = wl[sheet_name]

            # Obtiene la tabla de la hoja por su nombre
            tabla = wh.tables[table_name]
            # Obtiene la referencia actual de la tabla (ejemplo: 'A1:H10')
            ref_actual = tabla.ref
            # Extrae la coordenada final de la tabla (ejemplo: 'H10')
            coordenada_final = ref_actual.split(':')[-1]
            # Convierte la coordenada final en número de fila y columna
            fila_final_actual, columna_final_num = coordinate_to_tuple(coordenada_final)
            # Calcula la fila donde se insertarán los nuevos datos
            fila_inicio_nuevos_datos = fila_final_actual + 1
                
            # Inserta los nuevos datos fila por fila en la hoja
            for i, fila_nueva in enumerate(resumen):
                for j, valor in enumerate(fila_nueva):
                    wh.cell(row=fila_inicio_nuevos_datos + i, column=j + 1, value=valor)
            
            # Actualiza la referencia de la tabla para incluir las nuevas filas
            fila_final_nueva = fila_final_actual + len(resumen)
            columna_final_letra = get_column_letter(columna_final_num)
            referencia_inicial = ref_actual.split(':')[0]
            nueva_referencia = f'{referencia_inicial}:{columna_final_letra}{fila_final_nueva}'
            tabla.ref = nueva_referencia
            # --- Fin del código de openpyxl ---

            # Guarda el archivo modificado en un nuevo stream de bytes
            excel_stream_out = io.BytesIO()
            wl.save(excel_stream_out)
            excel_stream_out.seek(0)  # Mueve el cursor al inicio del stream
            
            # Sube el archivo modificado de vuelta a SharePoint
            success = sharepoint_client.upload_file(EXCEL_URL, excel_stream_out)

        except Exception as e:
            print(f"Error al procesar el archivo Excel: {e}")
            
    else:
        print("No se pudo descargar el archivo.")

def send_data(df, sheet, table, sharepoint_client):
    manual = df.values.tolist()
    if manual != []:
        modify_excel_file(manual, sheet, table, sharepoint_client)
