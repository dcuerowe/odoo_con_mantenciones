import xmlrpc.client
import json
import base64
import pandas as pd
import numpy as np
import re
from conn_sharepoint import get_auth_token, get_file_from_sharepoint, upload_file_to_sharepoint
from office365.sharepoint.client_context import ClientContext
from office365.sharepoint.files.file import File
from office365.runtime.auth.user_credential import UserCredential
from pathlib import Path
import openpyxl
import io
from io import BytesIO
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.utils import get_column_letter
import sqlite3
import schedule
from reportlab.lib.pagesizes import letter, A4, legal
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image
from reportlab.platypus.frames import Frame
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY, TA_LEFT, TA_RIGHT
from reportlab.platypus.flowables import HRFlowable
from reportlab.lib.utils import ImageReader
from datetime import datetime
from datetime import date
from zoneinfo import ZoneInfo
from datetime import timezone, timedelta
from dotenv import load_dotenv
import os
import requests
import time
import pytz
import locale
import tabulate
import traceback
import certifi

# #----------------------------------------------------------------------------------------
# #CLASES

class Sharepoint:

    def __init__(self):
        self.token = get_auth_token()


        
    def download_file(self, file_name, folder_name = ''):
        """
        Descarga un archivo desde una carpeta específica de SharePoint.

        Args:
            file_name (str): Nombre del archivo a descargar.
            folder_name (str): Nombre de la carpeta dentro de la biblioteca de documentos.

        Returns:
            bytes: El contenido del archivo en formato binario si la descarga es exitosa.
            None: Si ocurre un error durante la descarga.
        """

        try:
            # Obtiene la referencia al archivo usando la URL relativa
            file = get_file_from_sharepoint(file_name, self.token)
            # Devuelve el contenido binario del archivo
            return file.content
        
        except Exception as e:
            # Si ocurre un error, lo muestra y devuelve None
            print(f"Error al descargar el archivo {file_name} de SharePoint: {e}")
            return None
    
    def upload_file(self, file_name, content_stream, content_type = None, folder_name = ''):
        """Sube un archivo binario a una carpeta de SharePoint."""
        
        try:
            # Obtiene la referencia de la carpeta destino en SharePoint
        
            content_stream.seek(0)  # Asegura que el stream esté al inicio
            # Sube el archivo usando el contenido del stream
            target_file = upload_file_to_sharepoint(file_name, self.token, content_stream.getvalue(), content_type)

            # Ejecuta la consulta para completar la subida
            

            print(f"\n-> Archivo '{file_name}' subido con éxito en el intento 1")
            return True  # Retorna True si la subida fue exitosa
        
        except Exception as e:
            error_str = str(e)
            # Verifica si el error es por archivo bloqueado en SharePoint
            if "SPFileLockException" in error_str or "423 Client Error: Locked" in error_str:
                print(f"El archivo '{file_name}' está bloqueado. Reintentando en 5 segundos...")
                time.sleep(5)
            else:
                # Si es otro tipo de error, lo reporta y termina
                print(f"Error al subir el archivo a SharePoint: {error_str}")
                return None  # Retorna None en caso de error inesperado


# #----------------------------------------------------------------------------------------
# #FUNCIONES


def modify_excel_file(resumen, sheet_name, table_name):
    # Descarga el archivo Excel desde SharePoint
    excel = sp.download_file('https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04.%20Instalación%20y%20Mantenimiento/Trazabilidad%20de%20mantenciones%20y%20calibraciones/Captura.xlsx:/content')
    if excel:
        try:
            # Convierte los bytes descargados en un objeto BytesIO para manipulación en memoria
            excel_file = io.BytesIO(excel)
            # Carga el archivo Excel en openpyxl
            wl = openpyxl.load_workbook(excel_file)
            # Selecciona la hoja de trabajo especificada
            wh = wl[sheet_name]

            # Obtiene la tabla de la hoja por su nombre
            tabla = wh.tables[table_name]
            # Obtiene la referencia actual de la tabla (ejemplo: 'A1:H10')
            ref_actual = tabla.ref
            # Extrae la coordenada final de la tabla (ejemplo: 'H10')
            coordenada_final = ref_actual.split(':')[-1]
            # Convierte la coordenada final en número de fila y columna
            fila_final_actual, columna_final_num = coordinate_to_tuple(coordenada_final)
            # Calcula la fila donde se insertarán los nuevos datos
            fila_inicio_nuevos_datos = fila_final_actual + 1
                
            # Inserta los nuevos datos fila por fila en la hoja
            for i, fila_nueva in enumerate(resumen):
                for j, valor in enumerate(fila_nueva):
                    wh.cell(row=fila_inicio_nuevos_datos + i, column=j + 1, value=valor)
            
            # Actualiza la referencia de la tabla para incluir las nuevas filas
            fila_final_nueva = fila_final_actual + len(resumen)
            columna_final_letra = get_column_letter(columna_final_num)
            referencia_inicial = ref_actual.split(':')[0]
            nueva_referencia = f'{referencia_inicial}:{columna_final_letra}{fila_final_nueva}'
            tabla.ref = nueva_referencia
            # --- Fin del código de openpyxl ---

            # Guarda el archivo modificado en un nuevo stream de bytes
            excel_stream_out = io.BytesIO()
            wl.save(excel_stream_out)
            excel_stream_out.seek(0)  # Mueve el cursor al inicio del stream
            
            # Sube el archivo modificado de vuelta a SharePoint
            success = sp.upload_file('https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04.%20Instalación%20y%20Mantenimiento/Trazabilidad%20de%20mantenciones%20y%20calibraciones/Captura.xlsx:/content', excel_stream_out)

        except Exception as e:
            print(f"Error al procesar el archivo Excel: {e}")
            
    else:
        print("No se pudo descargar el archivo.")


def ordenar_respuestas(estructura, respuestas):
    # Mapeo de IDs de preguntas a títulos
    question_id_to_title = {q['questionId']: q['title'] for q in estructura['data']['questions']}

    # Lista de DataFrames para almacenar las respuestas
    dfs = []

    # Procesamiento de cada submission
    for submission in respuestas['data']['formSubmissions']:  # Itera sobre cada formulario enviado
        submission_data = {
            '#': submission['entryNum'],  # Guarda el número de entrada (OT)
            'user': submission['submittingUserId']  # Guarda el ID del usuario que envió el formulario
        }
        # Itera sobre cada respuesta dentro del formulario
        for answer in submission.get('answers', []):
            question_id = answer['questionId']  # Obtiene el ID de la pregunta
            # Busca el título de la pregunta usando el ID, si no lo encuentra usa un texto por defecto
            question_title = question_id_to_title.get(question_id, f"Pregunta {question_id}")

            value = None  # Inicializa el valor de la respuesta

            # Solo considerar respuestas que no estén vacías ni ocultas
            if not answer.get('wasSubmittedEpmty', False) and not answer.get('wasHidden', False):
                question_type = answer.get('questionType', 'unknown')  # Obtiene el tipo de pregunta

                # Procesa la respuesta según el tipo de pregunta
                if question_type == 'openEnded':
                    value = answer.get('value', 'error')  # Respuesta de texto libre
                elif question_type == 'multipleChoice':
                    # Une las opciones seleccionadas en una cadena separada por comas
                    selected = [opt['text'] for opt in answer.get('selectedAnswers', [])]
                    value = ', '.join(selected) if selected else 'Ninguna respuesta seleccionada'
                elif question_type == 'yesNo':
                    value = answer.get('selectedIndex', '')  # Índice de respuesta sí/no
                elif question_type == 'datetime':
                    date_sub = answer.get('timestamp', '')  # Obtiene el timestamp
                    dt_utc = datetime.fromtimestamp(date_sub, tz=timezone.utc)  # Convierte a fecha UTC
                    value = dt_utc.strftime("%d/%m/%Y")  # Formatea la fecha
                elif question_type == 'description':
                    value = None  # Las descripciones no se almacenan
                elif question_type == 'image':
                    value = answer.get('images', '')  # Obtiene las imágenes adjuntas
                elif question_type == 'signature':
                    value = answer.get('images', '')  # Obtiene las firmas adjuntas
                elif question_type == 'rating':
                    value = answer.get('ratingValue', '')  # Obtiene el valor de la calificación
                else:
                    value = 'Tipo no reconocido'  # Tipo de pregunta no soportado

            # Asocia el valor procesado al título de la pregunta en el diccionario de respuestas
            submission_data[question_title] = value

        # Crea un DataFrame con los datos procesados de la submission
        df = pd.DataFrame([submission_data])
        dfs.append(df)  # Agrega el DataFrame a la lista

    return dfs  # Devuelve la lista de DataFrames
             

def filter_submissions(API_key_connecteam):
    # Define la zona horaria de Chile
    chile_tz = ZoneInfo("America/Santiago")

    # Obtiene la fecha actual
    today = date.today()   

    # Calcula la fecha de hace 5 días
    yesterday = today - timedelta(days=5)

    # Calcula el inicio del día (medianoche) de hace 5 días en la zona horaria de Chile
    start_of_day_chile = datetime.combine(yesterday, datetime.min.time(), tzinfo=chile_tz)
    # Convierte el inicio del día a UTC
    start_of_day_utc =  start_of_day_chile.astimezone(timezone.utc)
    # Obtiene el timestamp en segundos para el inicio del rango
    start_timestamp_ms = int(start_of_day_utc.timestamp())

    # Calcula el fin del día de mañana en la zona horaria de Chile
    end_of_day_chile = datetime.combine(today + timedelta(days=1), datetime.min.time(), tzinfo=chile_tz)
    # Convierte el fin del día a UTC
    end_of_day_utc = end_of_day_chile.astimezone(timezone.utc)
    # Obtiene el timestamp en segundos para el final del rango, restando 1 para incluir solo hasta el final de hoy
    end_timestamp_ms = int(end_of_day_utc.timestamp()) - 1

    # Construye la URL para la API de Connecteam con los parámetros de rango de fechas y paginación
    url = f"https://api.connecteam.com/forms/v1/forms/12914411/form-submissions?submittingStartTimestamp={start_timestamp_ms}&submittingEndTime={end_timestamp_ms}&limit=100&offset=0"

    # Define los encabezados para la solicitud HTTP, incluyendo la API key
    headers = {
        "accept": "application/json",
        "X-API-KEY": f"{API_key_connecteam}"
    }

    # Realiza la solicitud GET a la API de Connecteam
    response = requests.get(url, headers=headers)
    # Convierte la respuesta en formato JSON
    response_json = response.json()
    # Retorna el JSON con las submissions filtradas por fecha
    return response_json

def all_submission(API_key_connecteam):
    """
    Obtiene las últimas 20 respuestas de un formulario específico desde la API de Connecteam.
    Parámetros:
        API_key_connecteam (str): Clave API necesaria para autenticar la solicitud a Connecteam.
    Funcionamiento:
        1. Define la URL del endpoint de Connecteam para obtener las respuestas del formulario.
        2. Construye los headers de la solicitud, incluyendo el API key para autenticación.
        3. Realiza una solicitud GET a la API usando la URL y los headers definidos.
        4. Convierte la respuesta recibida en formato JSON.
        5. Retorna el JSON con los datos de las respuestas del formulario.
    Retorna:
        dict: Un diccionario con la información de las respuestas del formulario obtenidas desde la API.
    """
    url = "https://api.connecteam.com/forms/v1/forms/12914411/form-submissions?limit=20&offset=0"

    headers = {"accept": "application/json",
            "X-API-KEY": f"{API_key_connecteam}"}

    response = requests.get(url, headers=headers)
    response_json = response.json()
    return response_json    


def form_structure(API_key_connecteam):
    """
    Obtiene la estructura de un formulario específico desde la API de Connecteam.
    Args:
        API_key_connecteam (str): Clave API necesaria para autenticar la solicitud a Connecteam.
    Returns:
        dict: Respuesta en formato JSON que contiene la estructura del formulario solicitado.
    Funcionamiento:
        1. Define la URL del formulario específico que se desea consultar.
        2. Prepara los encabezados de la solicitud, incluyendo el tipo de respuesta y la clave API.
        3. Realiza una solicitud GET a la API de Connecteam usando la URL y los encabezados definidos.
        4. Convierte la respuesta recibida en formato JSON.
        5. Retorna el JSON con la estructura del formulario.
    """
    url = "https://api.connecteam.com/forms/v1/forms/12914411"

    headers = {"accept": "application/json",
            "X-API-KEY": f"{API_key_connecteam}"}

    response = requests.get(url, headers=headers)
    response_json = response.json()
    return response_json

def user(API_key_connecteam, user_id):
    """
    Obtiene el nombre completo de un usuario activo desde la API de Connecteam.
    Args:
        API_key_connecteam (str): Clave API para autenticar la solicitud a Connecteam.
        user_id (int): ID del usuario cuyo nombre se desea obtener.
    Returns:
        str: Nombre completo del usuario (primer nombre y apellido).
    Funcionamiento:
        1. Construye la URL de la API con los parámetros necesarios, incluyendo el ID del usuario y el estado activo.
        2. Define los headers para la solicitud, incluyendo el API key.
        3. Realiza una solicitud GET a la API de Connecteam.
        4. Convierte la respuesta en formato JSON.
        5. Extrae el primer nombre y apellido del usuario desde la respuesta JSON.
        6. Retorna el nombre completo del usuario.
    """
    url = f"https://api.connecteam.com/users/v1/users?limit=30&offset=0&order=asc&userIds={int(user_id)}&userStatus=active"

    headers = {
        "accept": "application/json",
        "X-API-KEY": f"{API_key_connecteam}"
    }

    response = requests.get(url, headers=headers)

    response_json = response.json()

    nombre_usuario = response_json['data']['users'][0]["firstName"] + " " + response_json['data']['users'][0]["lastName"]
    return nombre_usuario


try:
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
except:
    pass  # Si no está disponible, usar el locale por defecto

def crear_estilos_personalizados():
    """
    Crea y retorna una colección de estilos personalizados para usar en documentos PDF con ReportLab.
    - styles = getSampleStyleSheet(): Obtiene una hoja de estilos base de ReportLab.
    - styles.add(ParagraphStyle(...)): 
        Agrega un estilo personalizado para el título principal ('TituloPrincipal'), 
        basado en 'Heading1', con fuente grande, negrita, centrado y espacio adicional antes y después.
    - styles.add(ParagraphStyle(...)): 
        Agrega un estilo para subtítulos ('Subtitulo'), basado en 'Heading2', 
        con fuente mediana, negrita y espaciado adecuado.
    - styles.add(ParagraphStyle(...)): 
        Agrega un estilo para mostrar información de fecha ('InfoFecha'), 
        basado en 'Normal', con fuente pequeña, color gris, alineado a la derecha y espacio después.
    - styles.add(ParagraphStyle(...)): 
        Agrega un estilo para el texto de introducción ('Introduccion'), 
        basado en 'Normal', con fuente estándar, justificado y espaciado personalizado.
    - styles.add(ParagraphStyle(...)): 
        Agrega un estilo para el pie de página ('PiePagina'), 
        basado en 'Normal', con fuente pequeña, color gris y centrado.
    Returns:
        StyleSheet1: Colección de estilos personalizados para usar en la generación de PDFs.
    """

    styles = getSampleStyleSheet()
    
    # Estilo para el título principal
    styles.add(ParagraphStyle(
        name='TituloPrincipal',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.black,
        spaceAfter=20,
        spaceBefore=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    ))
    
    # Estilo para subtítulos
    styles.add(ParagraphStyle(
        name='Subtitulo',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.black,
        spaceAfter=12,
        spaceBefore=16,
        fontName='Helvetica-Bold'
    ))
    
    # Estilo para información de fecha
    styles.add(ParagraphStyle(
        name='InfoFecha',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_RIGHT,
        spaceAfter=20,
        fontName='Helvetica-Oblique'
    ))
    
    # Estilo para texto de introducción
    styles.add(ParagraphStyle(
        name='Introduccion',
        parent=styles['Normal'],
        fontSize=11,
        alignment=TA_JUSTIFY,
        spaceAfter=16,
        spaceBefore=8,
        leftIndent=0,
        rightIndent=0
    ))
    
    # Estilo para el pie de página
    styles.add(ParagraphStyle(
        name='PiePagina',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    ))
    
    return styles

def crear_tabla_profesional(dataframe_trabajo, styles):
    """
    Crea una tabla profesional en formato ReportLab a partir de los datos de un DataFrame.
    Args:
        dataframe_trabajo (pd.DataFrame): DataFrame que contiene los datos del trabajo profesional. 
            Se espera que la primera fila contenga los valores de los campos requeridos.
        styles (dict): Diccionario de estilos de ReportLab para formatear los textos de la tabla.
    Returns:
        Table: Objeto Table de ReportLab listo para ser insertado en un documento PDF.
    Funcionamiento:
        1. Extrae los campos relevantes de la primera fila del DataFrame y los almacena en un diccionario.
        2. Convierte el diccionario en un DataFrame para facilitar la manipulación y el formateo vertical.
        3. Transforma el DataFrame para que cada campo y su respuesta sean una fila.
        4. Prepara la lista de datos para la tabla, incluyendo los encabezados.
        5. Formatea cada celda como Paragraph para aplicar estilos y permitir texto enriquecido.
        6. Define los anchos de columna para una presentación óptima.
        7. Aplica un estilo personalizado a la tabla, incluyendo colores, fuentes, alineaciones y espaciado.
        8. Devuelve la tabla lista para ser utilizada en la generación de un PDF profesional.
    """

    campos = {
        'OT:': dataframe_trabajo.iloc[0,0],
        'Técnico:': dataframe_trabajo.iloc[0,1],
        'Proyecto:': dataframe_trabajo.iloc[0,2],
        'Fecha de realización:': dataframe_trabajo.iloc[0,3],
        'Cliente:': dataframe_trabajo.iloc[0,4],
        'Equipo/instrumento:': dataframe_trabajo.iloc[0,5],
        'Modelo:': dataframe_trabajo.iloc[0,6],
        'N° de serie:': dataframe_trabajo.iloc[0,7]

    }

    df_campos = pd.DataFrame(campos, index=[0])
    
    # Preparar datos para la tabla
    df_vertical = df_campos.T.reset_index()
    df_vertical.columns = ['Campo', 'Respuesta']
    
    # Crear encabezados de tabla
    headers = ['Campo', 'Respuesta']
    data_list = [headers]
    
    # Agregar datos con formato mejorado
    for _, row in df_vertical.iterrows():
        campo = str(row['Campo'])
        respuesta = str(row['Respuesta'])
        data_list.append([campo, respuesta])
    
    # Convertir a Paragraphs para mejor control del formato
    data_for_table = []
    for i, row in enumerate(data_list):
        if i == 0:  # Encabezados
            formatted_row = [Paragraph(f"<b>{str(cell)}</b>", styles['Normal']) for cell in row]
        else:  # Datos
            # Campo en negrita, respuesta normal
            campo_cell = Paragraph(f"<b>{str(row[0])}</b>", styles['Normal'])
            respuesta_cell = Paragraph(str(row[1]), styles['Normal'])
            formatted_row = [campo_cell, respuesta_cell]
        data_for_table.append(formatted_row)
    
    # Crear tabla con anchos optimizados
    table = Table(data_for_table, colWidths=[5*cm, 12*cm], repeatRows=1)
    

    COLOR_NARANJA_WETECHS = HexColor("#EA6500")
    table_style = TableStyle([

        # Encabezado
        ('BACKGROUND', (0, 0), (-1, 0), COLOR_NARANJA_WETECHS),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Filas de datos
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Bordes y espaciado
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
        
        # Alternar colores de fila para mejor legibilidad
        #('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ])
    
    table.setStyle(table_style)
    return table



def add_header_first_page(canvas, doc):
    # --- 1. Cargar el logo desde la URL ---
    # URL del logo de We-Techs
    LOGO_URL = "https://we-techs-static-bucket.s3.amazonaws.com/static/images/logo-middle.png"

    try:
        response = requests.get(LOGO_URL)
        response.raise_for_status() # Lanza un error si la descarga falla
        logo_bytes = io.BytesIO(response.content)
        # Define el tamaño del logo. Ajusta estos valores si es necesario.
        img_reader = ImageReader(logo_bytes)
        original_width, original_height = img_reader.getSize()

        # Define el ancho deseado en cm y calcula la altura proporcionalmente
        logo_width_cm = 2.5 * cm
        aspect_ratio = original_height / original_width
        logo_height_cm = logo_width_cm * aspect_ratio
    
    except Exception as e:
        print(f"Error al cargar el logo desde la URL: {e}")
        logo_bytes = None
    
    if logo_bytes:
        # Rebobinar el stream de bytes
        logo_bytes.seek(0)
        
        # *** CAMBIO CLAVE: USAR ImageReader ***
        # Pasamos el stream de bytes a ImageReader, que es el formato
        # que ReportLab espera para dibujar imágenes en memoria.
        logo_image = ImageReader(logo_bytes)
        
        x_pos = legal[0] - doc.rightMargin - logo_width_cm
        y_pos = legal[1] - doc.topMargin - logo_height_cm + 1.5 * cm
        
        # Ahora pasamos el objeto ImageReader a drawImage
        canvas.drawImage(
            logo_image,
            x=x_pos,
            y=y_pos,
            width=logo_width_cm,
            height=logo_height_cm,
            mask='auto'
        )

def informe_pdf_profesional(numero_visita, tipo_trabajo, dataframe_visita, dataframe_trabajo, equipo):
    """
    Genera un informe PDF profesional detallando los trabajos realizados en una visita técnica.
    Args:
        numero_visita (str/int): Identificador de la visita técnica.
        tipo_trabajo (str): Tipo de trabajo realizado ('MC' para mantención correctiva, 'MP' para mantención preventiva).
        dataframe_visita (pd.DataFrame): DataFrame con información general de la visita.
        dataframe_trabajo (pd.DataFrame): DataFrame con el detalle de los trabajos realizados.
        equipo (str): Nombre o identificador del equipo intervenido.
    Returns:
        io.BytesIO: Buffer en memoria que contiene el PDF generado.
    Funcionamiento:
        - Define el tipo de mantención según el código recibido.
        - Inicializa un buffer en memoria para almacenar el PDF.
        - Configura el documento PDF con márgenes y tamaño de página profesional (A4).
        - Obtiene estilos personalizados para los elementos del informe.
        - Construye la estructura del informe (story) agregando:
            - Título principal con el punto de monitoreo.
            - Línea divisoria decorativa.
            - Fecha de generación del informe.
            - Subtítulo y texto introductorio con el tipo de trabajo y número de OT.
            - Tabla principal con los datos de los trabajos realizados.
        - Busca y agrega observaciones técnicas o generales si existen en los DataFrames.
        - Añade una línea divisoria y pie de página con información de la OT y equipo.
        - Extrae URLs de imágenes asociadas a la visita y las agrega al informe si existen.
        - Construye el documento PDF y retorna el buffer listo para ser guardado o enviado.
    """
    id_tipo_mantención = {'MC': 'mantención correctiva',
                    'MP': 'mantención preventiva',
                    'I': 'instalación'}
    # Configuración del archivo
    
    buffer = io.BytesIO()

    # Usar A4
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=legal,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2.5*cm,
        bottomMargin=2*cm
    )
    
    # Obtener estilos personalizados
    styles = crear_estilos_personalizados()
    
    # Lista de elementos del documento
    story = []
    
    # ENCABEZADO DEL DOCUMENTO
    # Título principal
    punto_monitoreo = dataframe_visita.get(f'{numero_visita}.1 Punto de monitoreo', ['N/A'])[0]
    title = Paragraph(
        f"Informe de trabajos<br/>{punto_monitoreo}",
        styles['TituloPrincipal']
    )
    story.append(title)
    
    # Información del nodo y número de visita
    # Línea divisoria decorativa
    line = HRFlowable(width="100%", thickness=1, lineCap='round', color=colors.orange)
    story.append(line)
    story.append(Spacer(1, 12))
    
    # Fecha de generación
    try:
        current_date = datetime.now().strftime("%d de %B de %Y")
    except:
        current_date = datetime.now().strftime("%d de %m de %Y")
    
    date_text = Paragraph(
        f"Fecha de generación: {current_date}",
        styles['InfoFecha']
    )
    story.append(date_text)
    #story.append(Spacer(1, 20))
    
    # SECCIÓN DE INTRODUCCIÓN
    intro_subtitle = Paragraph("Detalle", styles['Subtitulo'])
    story.append(intro_subtitle)
    
    ot_number = dataframe_visita['#'][0]


    intro_text = Paragraph(
        f"Este documento presenta el detalle de las tareas de {id_tipo_mantención[tipo_trabajo]} "
        f"realizadas sobre el equipo/instrumento señalado en tabla. ",
        styles['Introduccion']

    )

    story.append(intro_text)
    story.append(Spacer(1, 20))
    
    # Tabla principal con datos
    report_table = crear_tabla_profesional(dataframe_trabajo, styles)
    story.append(report_table)
    story.append(Spacer(1, 30))
    
    observaciones_a_agregar = []

    # 1. Buscar y guardar observaciones técnicas en dataframe_trabajo
    obs_cols = [
        col for col in dataframe_trabajo.columns 
        if 'observación' in str(col).lower() or 'observaciones' in str(col).lower()
    ]
    for col in obs_cols:
        obs_text = dataframe_trabajo[col].iloc[0]
        if pd.notna(obs_text) and str(obs_text).strip() and str(obs_text).lower() != 'nan':
            observaciones_a_agregar.append({
                'type': "Observaciones al equipo",
                'text': str(obs_text)
            })
            # Si solo quieres la primera observación técnica encontrada, puedes usar 'break' aquí.
            break 

    # 2. Buscar y guardar observaciones generales en dataframe_visita
    res_cols = [col for col in dataframe_visita.columns if 'resolución' in str(col).lower()]
    for col in res_cols:
        obs_text = dataframe_visita[col].iloc[0]
        if pd.notna(obs_text) and str(obs_text).strip() and str(obs_text).lower() != 'nan':
            observaciones_a_agregar.append({
                'type': "Observaciones generales",
                'text': str(obs_text)
            })
            # Si solo quieres la primera observación general encontrada, puedes usar 'break' aquí.
            break

    # 3. Agregar todas las observaciones encontradas al informe
    if observaciones_a_agregar:
        for obs in observaciones_a_agregar:
            obs_subtitle = Paragraph(obs['type'], styles['Subtitulo'])
            story.append(obs_subtitle)
            obs_paragraph = Paragraph(f"• {obs['text']}", styles['Introduccion'])
            story.append(obs_paragraph)
            story.append(Spacer(1, 20))

    story.append(Spacer(1, 30))
    line2 = HRFlowable(width="100%", thickness=0.5, lineCap='round', color=colors.lightgrey)
    story.append(line2)
    story.append(Spacer(1, 12))
    
    
    footer_text = Paragraph(
        f"Documento generado automáticamente • OT-{ot_number} • "
        f"{dataframe_trabajo.iloc[0,2]}",
        styles['PiePagina']
    )
    story.append(footer_text)



    lista_imagenes = []
    imagenes = dataframe_visita[f'{numero_visita}.4 Fotos recinto'][0]
    for img in imagenes:
        lista_imagenes.append(img['url'])

    # Agregar imágenes desde URLs
    if lista_imagenes:
        agregar_imagenes_desde_url(story, lista_imagenes)

    doc.build(story, onFirstPage=add_header_first_page)

    buffer.seek(0)
    
    return buffer

def agregar_imagenes_desde_url(story, lista_urls):
    """
    Agrega imágenes a una historia de ReportLab a partir de una lista de URLs.

    Args:
        story (list): Lista de elementos que conforman el documento PDF (usualmente una lista de Flowables de ReportLab).
        lista_urls (list): Lista de cadenas de texto, cada una representando una URL de imagen.

    Funcionamiento:
        - Calcula el ancho máximo permitido para la imagen en la página A4, dejando márgenes de 4 cm.
        - Itera sobre cada URL en la lista:
            - Intenta descargar la imagen usando requests.
            - Verifica que la descarga fue exitosa.
            - Convierte el contenido descargado en un objeto BytesIO para manipulación en memoria.
            - Crea un ImageReader para obtener las dimensiones originales de la imagen.
            - Si el ancho de la imagen excede el máximo permitido, escala proporcionalmente el ancho y el alto.
            - Agrega un salto de página antes de cada imagen para que cada una aparezca en una página nueva.
            - Inserta la imagen escalada en la historia.
        - Si ocurre algún error durante el proceso, imprime un mensaje indicando la URL y el error encontrado.
    """
    max_width = A4[0] - 4*cm 

    max_height = A4[1] - 4 * cm # alto máximo dentro de la página
    for url in lista_urls:
        try:
            response = requests.get(url)
            response.raise_for_status()
            img_bytes = BytesIO(response.content)
            img_reader = ImageReader(img_bytes)
            img_width, img_height = img_reader.getSize()
            # Escalado proporcional
            scale = min(max_width / img_width, max_height / img_height, 1.0)  # nunca agrandar, solo reducir

            new_width = img_width * scale
            new_height = img_height * scale

            story.append(PageBreak())
            story.append(Image(img_bytes, width=new_width, height=new_height))
        except Exception as e:
            print(f"Error al agregar imagen desde {url}: {e}")


# Función auxiliar para mantener compatibilidad con tu código existente
def informe_pdf(numero_visita, tipo_trabajo, dataframe_visita, dataframe_trabajo, equipo):
    return informe_pdf_profesional(numero_visita, tipo_trabajo, dataframe_visita, dataframe_trabajo, equipo)

# Resumen de notificaciones
def detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo, modelo, serial, trabajo_id, mensaje):
    """
    Agrega detalles de una operación al diccionario 'resumen', almacenando información relevante en listas asociadas a cada clave.

    Args:
        resumen (dict): Diccionario donde se almacenan los detalles de la operación. Cada clave corresponde a un campo y su valor es una lista.
        ot (str): Número de orden de trabajo (OT) que se agrega a la lista 'OT'.
        tecnico (str): Nombre del técnico responsable, agregado a la lista 'Técnico'.
        fecha (str): Fecha de revisión, agregada a la lista 'Fecha de revisión'. Si la clave no existe, se crea.
        proyecto (str): Nombre o identificador del proyecto, agregado a la lista 'Proyecto'.
        punto (str): Punto de monitoreo, agregado a la lista 'Punto de monitoreo'.
        tipo (str): Tipo de equipo/instrumento, agregado a la lista 'Equipo/instrumento'.
        modelo (str): Modelo del equipo/instrumento, agregado a la lista 'Modelo'.
        serial (str): Número de serie del equipo/instrumento, agregado a la lista 'N° serie'.
        trabajo_id (str): Identificador del tipo de trabajo realizado, agregado a la lista 'Tipo'.
        mensaje (str): Mensaje o comentario asociado a la operación, agregado a la lista 'Mensaje'.

    Funcionamiento:
        - Agrega el valor de 'ot' a la lista bajo la clave 'OT'.
        - Agrega el valor de 'tecnico' a la lista bajo la clave 'Técnico'.
        - Si la clave 'Fecha de revisión' no existe en 'resumen', la crea con una lista vacía y luego agrega 'fecha'.
        - Agrega el valor de 'proyecto' a la lista bajo la clave 'Proyecto'.
        - Agrega el valor de 'punto' a la lista bajo la clave 'Punto de monitoreo'.
        - Agrega el valor de 'modelo' a la lista bajo la clave 'Modelo'.
        - Agrega el valor de 'serial' a la lista bajo la clave 'N° serie'.
        - Agrega el valor de 'trabajo_id' a la lista bajo la clave 'Tipo'.
        - Agrega el valor de 'tipo' a la lista bajo la clave 'Equipo/instrumento'.
        - Agrega el valor de 'mensaje' a la lista bajo la clave 'Mensaje'.
    """
    resumen['OT'].append(ot)
    resumen['Técnico'].append(tecnico)
    resumen.setdefault('Fecha de revisión', []).append(fecha)
    resumen['Proyecto'].append(proyecto)
    resumen['Punto de monitoreo'].append(punto)
    resumen['Modelo'].append(modelo)
    resumen['N° serie'].append(serial)
    resumen['Tipo'].append(trabajo_id)
    resumen['Equipo/instrumento'].append(tipo)
    resumen['Mensaje'].append(mensaje)



#-----------------------------------------------------------------------------------------
#PROCESAMIENTO DE INFORMACIÓN
#Analizando cada sub que esta en forma de dataframe
def process_entrys(ordered_responses, API_key_c, resumen, exito):

    for df in ordered_responses:

        df = df.astype({'user': str}) #Eliminando las columnas que no se usaron
        
        df_con_datos = df.dropna(axis=1, how ='all') #Eliminando las columnas que no se usaron
        df_columnas = df_con_datos.columns.to_list() #Lista de columnas que si tienen datos

        index_user = df_con_datos.columns.get_loc('user')

        try:
            user_name = user(API_key_c, df_con_datos['user'][0])
        except Exception as e:
            user_name = "Usuario no encontrado"
            print(f"Error al obtener el nombre del usuario: {e}")
            traceback.print_exc()
        
        try:
            df_con_datos.iloc[0, index_user] = user_name # Añadir el nombre del usuario al DataFrame
        except Exception as e:
            print(f"Error al asignar el nombre del usuario al DataFrame: {e}")
            traceback.print_exc()


        #Elementos globales
        id_tipo_de_trabajo = ['MP', 'MC', 'I']
        
        id_mantencion = {'MC': 'corrective',
                        'MP': 'preventive'}
        
        intalaciones_interes = ['Tablero', 'Caudalímetro', 'Sensor de nivel', 'Sonda multiparamétrica', 'Otro']

        #Puntos que efectivamente se visitaron
        numeros_visita = set()
        for col in df_columnas:
            # Verificamos si el nombre de la columna comienza con un dígito
            if col and col[0].isdigit():
                # Extraemos el primer carácter (el número)
                numeros_visita.add(col[0])

        numeros_visita = sorted(list(numeros_visita))

        for i in numeros_visita:

            #Separación de los trabajos realizados
            try: 
                tipos_realizados = [tipo.strip() for tipo in df_con_datos[f'{i}.2 Tipo de trabajo a realizar'][0].split(',') ]
            except:
                tipos_realizados = df_con_datos[f'{i}.2 Tipo de trabajo a realizar']

            # Columnas del punto {1} | general
            columnas_visita = [columna for columna in df_columnas if columna.startswith(i)]
            #columnas_visita.append(f'{i} Proyecto') 
            columnas_visita = ['#', 'user', 'Fecha visita ', 'Nombre del Cliente'] + columnas_visita 

            #Dejando un dataframe a nivel de visita de punto
            df_visita = df_con_datos[columnas_visita].copy()


            #Validando si el punto se encuentra seteadao en el listado de connecteam
            if df_visita[f'{i}.1 Punto de monitoreo'][0] == "No encontrado":
            
                #Creamos la columna proyecto
                try:
                    index_columna_punto_proyecto = df_visita.columns.get_loc(f'{i} Proyecto')
                    df_visita.loc[:, f"{i}.1 Proyecto"] = df_visita.iloc[0, index_columna_punto_proyecto]
                    
                    #Definimos el punto ingresado manaualmente como el verdadero
                    index_columna_punto_no = df_visita.columns.get_loc(f'{i}.1 Punto de monitoreo')
                    index_columna_punto_si = df_visita.columns.get_loc(f'{i}.1 Indicar nombre del punto')

                    df_visita.iloc[0, index_columna_punto_no] = df_visita.iloc[0, index_columna_punto_si]

                    del df_visita[f'{i}.1 Indicar nombre del punto']
                except Exception as e:
                    print(f"Error al procesar el punto de monitoreo en OT {df_visita['#'][0]}: {e}")
                    # Si no se encuentra la columna, asignamos un valor por defecto
                    df_visita.loc[:, f"{i}.1 Proyecto"] = "Proyecto no especificado"
                    df_visita.loc[:, f"{i}.1 Punto de monitoreo"] = "Punto no especificado"
                    continue
                    

            else:
                #Buscando el indice de columna
                index_columna_punto = df_visita.columns.get_loc(f'{i}.1 Punto de monitoreo')

                #Definiendo el nombre del proyecto
                match = re.search(r"\[([^\]]*)\]", df_visita.iloc[0, index_columna_punto])
                if match:
                    df_visita.loc[:, f"{i}.1 Proyecto"] = match.group(1)
                else:
                    df_visita.loc[:, f"{i}.1 Proyecto"] = "Proyecto no especificado"
                                
                #Definiedo el nombre del punto
                df_visita.iloc[0, index_columna_punto] = re.sub(r"\[[^\]]*\]", "", df_visita.iloc[0, index_columna_punto]).strip() #Eliminando el nombre del proyecto

            #Definimos los ID de los tipos de trabajo realizados
            id_tipos_realizados = [item.split(' |')[0] for item in tipos_realizados]
            #id_tipos_realizados

            #Definimos los ID de tipos de trabajo de interes
            id_tipos_interes = []
            for tipo in id_tipos_realizados:
                if tipo in id_tipo_de_trabajo:
                    id_tipos_interes.append(tipo)
            id_tipos_interes #[MC, MP]
            
            #Cantidad de MP realizadas
            MP_prefijo = set()
            for col in df_visita.columns:
                if ' MP |' in col: # Buscamos ' MP |' para identificar las columnas de MP
                    # Extraemos el prefijo como '1.2.1 MP' o '1.2.2 MP'
                    prefix_end_index = col.find(' MP |') + 4 # Sumamos 4 para incluir ' MP'
                    prefix = col[:prefix_end_index].strip()
                    MP_prefijo.add(prefix)
            
            conteo_instancias_MP = len(MP_prefijo)

            #Cantidad de MC realizadas
            MC_prefijo = set()
            for col in df_visita.columns:
                if ' MC |' in col: # Buscamos ' MC |' para identificar las columnas de MC
                    # Extraemos el prefijo como '1.2.1 MC' o '1.2.2 MC'
                    prefix_end_index = col.find(' MC |') + 4 # Sumamos 4 para incluir ' MC'
                    prefix = col[:prefix_end_index].strip()
                    MC_prefijo.add(prefix)
            
            conteo_instancias_MC = len(MC_prefijo)

            #Cantidad de I realizadas
            I_prefijo = set()
            for col in df_visita.columns:
                if ' I |' in col: # Buscamos ' MP |' para identificar las columnas de MP
                    # Extraemos el prefijo como '1.2.1 MP' o '1.2.2 MP'
                    I_prefix_end_index = col.find(' I |') + 4 # Sumamos 4 para incluir ' MP'
                    I_prefix = col[:I_prefix_end_index].strip()
                    I_prefijo.add(I_prefix)
            
            conteo_instancias_I = len(I_prefijo)
            #print(df_visita)


            for id in id_tipos_realizados:
                #Iniciamos la filtración por tipos de trabajo
                columnas_trabajo = [columna for columna in df_visita.columns if f'{id}' in columna]
                columnas_trabajo = ['#', 'user', f"{i}.1 Proyecto", 'Fecha visita ', 'Nombre del Cliente'] + columnas_trabajo
                df_trabajo = df_visita[columnas_trabajo]

                proyecto = df_visita[f"{i}.1 Proyecto"][0]
                punto = df_visita[f'{i}.1 Punto de monitoreo'][0]
                ot = df_visita['#'][0]
                fecha = df_visita['Fecha visita '][0]
                tecnico = df_visita['user'][0]
                
                #Tratamiento para Mantención correctiva
                if id == "MC":
                    for equipo in range(1, conteo_instancias_MC+1):
                        filtro_MC = f"{i}.2.{equipo} MC"        
                        columnas_equipo_MC = df_trabajo.filter(like=filtro_MC).columns.to_list()
                        columnas_equipo_MC = ['#', 'user', f"{i}.1 Proyecto", 'Fecha visita ', 'Nombre del Cliente'] + columnas_equipo_MC
                        df_trabajo_equipo_MC = df_trabajo[columnas_equipo_MC]
                        dic_trabajo_MC = df_trabajo_equipo_MC.to_dict(orient='records')[0]
                        modelo_MC = dic_trabajo_MC[f"{i}.2.{equipo} MC | Modelo"]
                        tipo_MC = dic_trabajo_MC[f"{i}.2.{equipo} MC | ¿A qué se le realiza mantenimiento correctivo?"]
                        serial_MC = dic_trabajo_MC[f'{i}.2.{equipo} MC | N° de serie']
    
                        
                        #Asegurando que el serial pase de float a int
                        for llave, valor in dic_trabajo_MC.items():
                                    if isinstance(valor,float):
                                        dic_trabajo_MC[llave] = int(valor)


                        pdf_stream_MC = informe_pdf(i, id, df_visita, df_trabajo_equipo_MC, equipo)
                        
                        nombre_archivo_MC = f"informe_OT-{df_visita['#'][0]}_{i}_{id}_{equipo}.pdf"

                        pdf_stream_MC.seek(0)

                        try:
                            contenido_pdf = pdf_stream_MC.read()
                            informe_codficado_MC = base64.b64encode(contenido_pdf).decode('utf-8')
                        except FileNotFoundError:
                            exit()


                        #------------------------------------------------------------------------
                        #Busqueda del ID del equipo en la base de datos maintenance.equipment
                        try:
                            equipment_MC = models.execute_kw(db, uid, password,
                                'maintenance.equipment', 'search_read',
                                [[
                                    ['serial_no', '=', serial_MC]
                                ]],
                                {'limit': 1})

                            if equipment_MC:

                                #Validamos la ubicación del equipo:
                                id_number_MC = equipment_MC[0]['id']
                                
                                if equipment_MC[0]['x_studio_location']:
                                    location_MC = equipment_MC[0]['x_studio_location'][1]
                                else:
                                    location_MC = False
                                

                                if location_MC != df_con_datos[f'{i}.1 Punto de monitoreo'][0]:
                                    detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MC, modelo_MC, serial_MC, id, 
                                                f'La ubicación indicada en la OT ({df_con_datos[f"{i}.1 Punto de monitoreo"][0]}) es distinta a la registrada en Odoo ({location_MC}). Revisar OT.')
                                    
                                    try:
                                        new_location_MC = models.execute_kw(db, uid, password,
                                            'maintenance.equipment', 'message_post', [
                                                id_number_MC,
                                            ], {
                                                'body': f"<p>La ubicación a cambiado.</p><p>Nueva ubicación: {punto}</p>",
                                                'message_type': 'comment',
                                                'subtype_xmlid': 'mail.mt_note',  # nota interna
                                            })
                                    except Exception as e:
                                        print(f'Error al notificar la nueva ubicación del equipo en Odoo: {e}')  
            
                                else:
                                    try:
                                        last_location_MC = models.execute_kw(db, uid, password,
                                            'maintenance.equipment', 'message_post', [
                                                id_number_MC,
                                            ], {
                                                'body': f"<p>Última ubicación: {punto}</p>",
                                                'message_type': 'comment',
                                                'subtype_xmlid': 'mail.mt_note',  # nota interna
                                            })
                                    except Exception as e:
                                        print(f'Error al notificar la ubicación del equipo en Odoo: {e}')


                                #Busqueda de solicitues correctivas que mantiene el equipo
                                
                                try:
                                    domain_filter_MC = [['equipment_id', '=', id_number_MC],
                                                    ['maintenance_type', '=', 'corrective']]

                                    request_ids_MC = models.execute_kw(db, uid, password,
                                                    'maintenance.request', 'search',
                                                    [domain_filter_MC])
                                    
                                    #Validación de si se programo o no el trabajo
                                    interruptor_MC = True    
                                    for ids_MC in request_ids_MC:
                                        campos_de_interes_MC = ['schedule_date', 'stage_id', "name"]
                                        try:
                                            request_data_MC = models.execute_kw(db, uid, password,
                                                'maintenance.request', 'read',
                                                [[ids_MC]],
                                                {'fields': campos_de_interes_MC})
                                            
                                            stage_id_MC = request_data_MC[0].get('stage_id') 
                                            schedule_date_MC = request_data_MC[0].get('schedule_date')

                                            if schedule_date_MC == False:
                                                continue
                                            elif schedule_date_MC != False and stage_id_MC[0] == 5:
                                                continue
                                            else:
                                                interruptor_MC = False
                                                break
                                        except Exception as e:
                                            print(e)

                                    #------------------------------------------------------------------------

                                    #Creación de request
                                    if interruptor_MC:
                                        try:
                                            fields_values_OT_MC = {
                                                'name': f"MANTENIMIENTO CORRECTIVO | {tipo_MC} {modelo_MC}",
                                                'equipment_id': id_number_MC, #Aquí debemos usar el ID númerico de la sonda
                                                'stage_id': '5', # 5 es finalizado 
                                                'maintenance_type': id_mantencion[id],
                                                'description': f"{dic_trabajo_MC[f'{i}.2.{equipo} MC | Observaciones']}",
                                                'schedule_date': f"{dic_trabajo_MC[f'Fecha visita ']}",
                                                'x_studio_informe': informe_codficado_MC
                                                #'x_studio_nmero_de_ot_1': f"{dic_trabajo_MC['#']}"
                                            }
                                            created_request_MC = models.execute_kw(db, uid, password,
                                                'maintenance.request',
                                                'create',
                                                [fields_values_OT_MC]
                                                )
                                            
                                            #Eliminación del informe PDF
                                            #os.remove(pdf_file)


                                            #Resgistro en resumen
                                            detalle_op(exito, ot, tecnico, fecha, proyecto, punto, tipo_MC, modelo_MC, serial_MC, id,
                                                        f'Se crea con éxito el registro de mantenimiento')
                                        
                                            try:    
                                                #Actualización de la actividad por defecto 'Maintenance Request'
                                                #Buscando el ID de la OT creada
                                                OT_id_MC = models.execute_kw(db, uid, password, #Esto nos entrega una lista
                                                    'maintenance.request', 'search',
                                                    [[
                                                        ['x_studio_nmero_de_ot_1', '=', fields_values_OT_MC['x_studio_nmero_de_ot_1']] 
                                                    ]],
                                                    {'limit': 1})

                                                if OT_id_MC:
                                                    OT_number_MC = OT_id_MC[0]

                                                    try:
                                                        #Buscamos el ID de la actividad existente para la OT_number
                                                        actividad_id_MC = models.execute_kw(db, uid, password,
                                                            'mail.activity', 'search_read',
                                                            [[
                                                                ['res_model', '=', 'maintenance.request'],
                                                                ['res_id', '=', OT_number_MC] 
                                                            ]],
                                                            {'limit': 1 })

                                                        #Actualizando actividad
                                                        try:
                                                            actividad_number_MC = actividad_id_MC[0]['id']
                                                            models.execute_kw(db, uid, password,
                                                                'mail.activity', 'action_feedback',
                                                                [
                                                                    [actividad_number_MC] # Lista de IDs de actividades a marcar como hechas
                                                                ],
                                                                {'feedback': f"Se ha completado desde API | Última ubicación: {df_visita[f'{i}.1 Punto de monitoreo'][0]}"}
                                                                ) # Mensaje opcional para el historial)

                                                        except Exception as e:
                                                            print(f"Error al actualizar la actividad de mantenimiento asociada: {e}")
                                                            continue        
                                                        
                                                    except Exception as e:
                                                        print(f"Error al buscar la actividad de manteniminto asociada: {e}") 
                                                        continue   
                                                else:
                                                    print(f"No se encontró la OT creada en Odoo para la solicitud de mantenimiento: {fields_values_OT_MC['x_studio_nmero_de_ot_1']}")
                                                    continue
                                            except Exception as e:
                                                print(f"Error al buscar request de mantenimiento recien creada: {e}")
                                                continue

                                            #Actualización de bitácora

                                        except Exception as e:
                                            print(f"Error al crear request MC para la OT-{dic_trabajo_MC['#']} en Odoo: {type(e)}")
                                            continue

                                    #------------------------------------------------------------------------
                                    #Actualización del request encontrado
                                    else:
                                        try:
                                            #Actualizando su estado a Finalizado
                                            update_MC = {
                                                'stage_id': 5,
                                                'x_studio_informe': informe_codficado_MC,
                                                'x_studio_nmero_de_ot_1': f"{dic_trabajo_MC['#']}"
                                            }

                                            update_stage_MC = models.execute_kw(
                                                db, uid, password,
                                                'maintenance.request',
                                                'write',
                                                [
                                                    [ids_MC], 
                                                    update_MC
                                                ],)
                                            
                                            if update_stage_MC:

                                                #os.remove(pdf_file) #Eliminación del informe PDF

                                                #Confirmación de validación
                                                detalle_op(exito, ot, tecnico, fecha, proyecto, punto, tipo_MC, modelo_MC, serial_MC, id, 
                                                            f'Se registra con éxito el mantenimiento correctivo programado: {request_data_MC[0]["name"]}')
            
                        
                                                #Actualiación de Bitácora

                                                #Buscamos el ID de la actividad existente para la OT
                                                try:
                                                    actividad_id_MC = models.execute_kw(db, uid, password,
                                                        'mail.activity', 'search_read',
                                                        [[
                                                            ['res_model', '=', 'maintenance.request'],
                                                            ['res_id', '=', ids_MC] 
                                                        ]],
                                                        {'limit': 1 })

                                                    #Actualizando actividad
                                                    if actividad_id_MC:

                                                        try:
                                                            actividad_number_MC = actividad_id_MC[0]['id']
                                                            models.execute_kw(db, uid, password,
                                                                'mail.activity', 'action_feedback',
                                                                [
                                                                    [actividad_number_MC] # Lista de IDs de actividades a marcar como hechas
                                                                ],
                                                                {'feedback': f"Se ha completado desde API | Última ubicación: {df_visita[f'{i}.1 Punto de monitoreo'][0]}"}) # Mensaje opcional para el historial)
                                                            
                                                        except Exception as e:
                                                            print(f"Error al actualizar estado de la actividad de mantenimiento: {e}")
                                                            continue
                                                                
                                                except Exception as e:
                                                    print(f"Error al listar de la actividad de mantenimiento: {e}")
                                                        
                                        except Exception as e:
                                            print(f"Error al actualizar estado de solicitud de mantenimiento: {e}")  
                                
                                except Exception as e:
                                    print(f"Error al buscar las solicitudes de mantenimiento: {e}")
                                    continue
                                
                            else:
                                try:
                                    sp.upload_file(f'https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04. Instalación y Mantenimiento/Trazabilidad de mantenciones y calibraciones/Informes de mantención/{nombre_archivo_MC}:/content', pdf_stream_MC, "application/pdf" )
                                except Exception as e:
                                    print(f"Error al subir el informe al Sharepoint: {e}")

                                detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MC, modelo_MC, serial_MC, id, 
                                            f'N° de serie no encontrado en Odoo. Revisar OT | {nombre_archivo_MC}')
                                continue
                        except Exception as e:
                            print(f"Error al buscar equipo en base de Odoo MC: {e}")
                            continue

                #Tratamiento para mantención preventiva  
                elif id == "MP":
                    for equipo in range(1,conteo_instancias_MP+1):
                        filtro_MP = f"{i}.2.{equipo} MP"
                        
                        columnas_equipo_MP = df_trabajo.filter(like=filtro_MP).columns.to_list()
                        columnas_equipo_MP = ['#', 'user', f"{i}.1 Proyecto", 'Fecha visita ', 'Nombre del Cliente'] + columnas_equipo_MP

                        df_trabajo_equipo_MP = df_trabajo[columnas_equipo_MP]

                        dic_trabajo_MP = df_trabajo_equipo_MP.to_dict(orient='records')[0]

                        for llave, valor in dic_trabajo_MP.items():
                            if isinstance(valor,float):
                                dic_trabajo_MP[llave] = int(valor)

                        modelo_MP = dic_trabajo_MP[f"{i}.2.{equipo} MP | Modelo"] 
                        tipo_MP = dic_trabajo_MP[f"{i}.2.{equipo} MP | ¿A qué se le realiza mantenimiento preventivo?"]
                        ot_mp = dic_trabajo_MP['#']
                        fecha_mp = dic_trabajo_MP['Fecha visita ']


                        #print(df_trabajo_equipo_MP)
                        # CREACIÓN DE INFORME
                        pdf_stream_MP = informe_pdf(i, id, df_visita, df_trabajo_equipo_MP, equipo)
                        
                        nombre_archivo_MP = f"informe_OT-{df_visita['#'][0]}_{i}_{id}_{equipo}.pdf"

                        pdf_stream_MP.seek(0)


                        try:
                            contenido_pdf_MP = pdf_stream_MP.read()
                            informe_codificado_MP = base64.b64encode(contenido_pdf_MP).decode('utf-8')
                        except FileNotFoundError:
                            exit()

                        with open(nombre_archivo_MP, 'wb') as f:
                            f.write(contenido_pdf_MP)


                        #ACTUALIZACIÓN DE REQUEST
                        
                        #Buscamos las request que existan para el equipo en cuestión
                        serial_MP = dic_trabajo_MP[f'{i}.2.{equipo} MP | N° de serie']
                        try:
                            equipment_MP = models.execute_kw(db, uid, password,
                                'maintenance.equipment', 'search_read',
                                [[
                                    ['serial_no', '=', serial_MP] #Bucamos el id de tabla del instrumento dentro de 'maintenance.equipment'
                                ]],
                                {'limit': 1}
                                )
                        
                            if equipment_MP:
                                number_equipment_MP = equipment_MP[0]['id']

                                #Validación de ubicación
                                if equipment_MP[0]['x_studio_location']:
                                    location_MP = equipment_MP[0]['x_studio_location'][1]
                                else:
                                    location_MP = False
                                

                                if location_MP != df_con_datos[f'{i}.1 Punto de monitoreo'][0]:
                                    detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MP, modelo_MP, serial_MP, id, 
                                    f'La ubicación indicada en la OT ({df_con_datos[f"{i}.1 Punto de monitoreo"][0]}) es distinta a la registrada en Odoo ({punto}). Revisar OT.')
                                        
                                    try:
                                        new_location_MP = models.execute_kw(db, uid, password,
                                            'maintenance.equipment', 'message_post', [
                                                number_equipment_MP,
                                            ], {
                                                'body': f"<p>La ubicación a cambiado.</p><p>Nueva ubicación: {punto}</p>",
                                                'message_type': 'comment',
                                                'subtype_xmlid': 'mail.mt_note',  # nota interna
                                            })
                                    except Exception as e:
                                        print(f'Error al notificar la nueva ubicación del equipo en Odoo: {e}')                                    
                                    
                                else:
                                    try:
                                        last_location_MP = models.execute_kw(db, uid, password,
                                            'maintenance.equipment', 'message_post', [
                                                number_equipment_MP,
                                            ], {
                                                'body': f"<p>Última ubicación: {punto}</p>",
                                                'message_type': 'comment',
                                                'subtype_xmlid': 'mail.mt_note',  # nota interna
                                            })
                                    except Exception as e:
                                        print(f'Error al notificar la ubicación del equipo en Odoo: {e}')

                                try:
                                    domain_filter = [['equipment_id', '=', number_equipment_MP],
                                                    ['maintenance_type', '=', 'preventive']]

                                    request_ids_MP = models.execute_kw(db, uid, password,
                                                    'maintenance.request', 'search',
                                                    [domain_filter])
                                    
                                    #Iterando sobre las solicitudes que tiene el equipo
                                    if request_ids_MP:
                                        interruptor_MP = True    
                                        for ids_MP in request_ids_MP:
                                            campos_de_interes_MP = ['schedule_date', 'stage_id', 'name']
                                            try:
                                                request_data_MP = models.execute_kw(db, uid, password,
                                                    'maintenance.request', 'read',
                                                    [[ids_MP]],
                                                    {'fields': campos_de_interes_MP})
                                                
                                                stage_id_MP = request_data_MP[0].get('stage_id') 
                                                schedule_date_MP = request_data_MP[0].get('schedule_date')

                                                if schedule_date_MP == False:
                                                    continue
                                                elif schedule_date_MP != False and stage_id_MP[1] == "Finalizado":
                                                    continue
                                                else:
                                                    interruptor_MP = False
                                                    break
                                            except Exception as e:
                                                print(e)
                                        if interruptor_MP == False:
                                            try:
                                                #Atualizando su estado a Finalizado
                                                update_MP = {
                                                    'stage_id': 5,
                                                    'x_studio_informe': informe_codificado_MP,
                                                    'x_studio_nmero_de_ot_1': f"{dic_trabajo_MP['#']}"
                                                }

                                                update_stage_MP = models.execute_kw(
                                                    db, uid, password,
                                                    'maintenance.request',
                                                    'write',
                                                    [
                                                        [ids_MP], 
                                                        update_MP
                                                    ],)
                                                
                                                if update_stage_MP:

                                                    detalle_op(exito, ot, tecnico, fecha, proyecto, punto, tipo_MP, modelo_MP, serial_MP, id, 
                                                                f'Se registra con exito el mantenimiento preventivo programado: {request_data_MP[0].get("name")}')
                                                    
                                                    #Actualización de bitácora
                                                    try:
                                                        actividad_id_MP = models.execute_kw(db, uid, password,
                                                            'mail.activity', 'search_read',
                                                            [[
                                                                ['res_model', '=', 'maintenance.request'],
                                                                ['res_id', '=', ids_MP] 
                                                            ]],
                                                            {'limit': 1 })

                                                        #Actualizando actividad
                                                        if actividad_id_MP:
                                                            try:
                                                                actividad_number_MP = actividad_id_MP[0]['id']
                                                                models.execute_kw(db, uid, password,
                                                                    'mail.activity', 'action_feedback',
                                                                    [
                                                                        [actividad_number_MP] # Lista de IDs de actividades a marcar como hechas
                                                                    ],
                                                                    {'feedback': f"Se ha completado desde API | Última ubicación: {df_visita[f'{i}.1 Punto de monitoreo'][0]}"}) # Mensaje opcional para el historial)
                                                                
                                                            except Exception as e:
                                                                print(f"Error al actualizar estado de la actividad de mantenimiento: {e}")

                                                    except Exception as e:
                                                        print(f"Error al listar de la actividad de mantenimiento: {e}")
                                            
                                                    #  Buscamos el ID de la actividad existente para la OT
                                                            
                                            except Exception as e:
                                                print(f"Error al actualizar estado de solicitud de mantenimiento MP: {e}")
                                                traceback.print_exc()
                                                continue

                                        else:

                                            try:
                                                sp.upload_file(f'https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04. Instalación y Mantenimiento/Trazabilidad de mantenciones y calibraciones/Informes de mantención/{nombre_archivo_MP}:/content', pdf_stream_MP, "application/pdf" )

                                            except Exception as e:
                                                print(f"Error al subir el informe al Sharepoint: {e}")

                                            detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MP, modelo_MP, serial_MP, id, 
                                                        f'El equipo/instrumento no cuenta con una solicitud de mantenimiento preventivo programada. Revisar OT | {nombre_archivo_MP}')

                                    else:

                                        try:
                                            sp.upload_file(f'https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04. Instalación y Mantenimiento/Trazabilidad de mantenciones y calibraciones/Informes de mantención/{nombre_archivo_MP}:/content', pdf_stream_MP, "application/pdf" )
                                        except Exception as e:
                                            print(f"Error al subir el informe al Sharepoint: {e}")

                                        detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MP, modelo_MP, serial_MP, id, 
                                                    f'El equipo no tiene un plan de mantenimiento cargado en Odoo. Revisar OT | {nombre_archivo_MP}')
        
                                except Exception as e: 
                                    print(f"Error al obtener información de la solicitudes de mantenimiento: {e}")
                                    continue 

                            else:   

                                try:
                                    sp.upload_file(f'https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04. Instalación y Mantenimiento/Trazabilidad de mantenciones y calibraciones/Informes de mantención/{nombre_archivo_MP}:/content', pdf_stream_MP, "application/pdf" )
                                except Exception as e:
                                    print(f"Error al subir el informe al Sharepoint: {e}")

                                detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_MP, modelo_MP, serial_MP, id, 
                                            f'N° de serie no encontrado en Odoo. Revisar OT | {nombre_archivo_MP}')
                                continue
                    
                        except Exception as e:
                                    print(f"Error al buscar equipo en base de Odoo MP: {e}")
                
                #Tratamiento para Instalaciones
                elif id == "I":
                    for equipo in range(1, conteo_instancias_I+1):
                        filtro_I = f"{i}.2.{equipo} I"        
                        columnas_equipo_I = df_trabajo.filter(like=filtro_I).columns.to_list()
                        columnas_equipo_I = ['#', 'user', f"{i}.1 Proyecto", 'Fecha visita ', 'Nombre del Cliente'] + columnas_equipo_I
                        df_trabajo_equipo_I = df_trabajo[columnas_equipo_I]
                        dic_trabajo_I = df_trabajo_equipo_I.to_dict(orient='records')[0]
                
                        modelo_I = dic_trabajo_I[f"{i}.2.{equipo} I | Modelo"]
                        tipo_I = dic_trabajo_I[f"{i}.2.{equipo} I | Tipo de equipo/instrumento a instalar"]


                        pdf_stream_I = informe_pdf(i, id, df_visita, df_trabajo_equipo_I, equipo)

                        
                        nombre_archivo_I = f"informe_OT-{df_visita['#'][0]}_{i}_{id}_{equipo}.pdf"

                        pdf_stream_I.seek(0)

                        # with open(nombre_archivo_I, 'wb') as f:
                        #     f.write(contenido_pdf_I)
                        # # El archivo 'nombre_archivo_I' se habrá creado en el mismo directorio donde se ejecuta el script.
                        # print(f"Informe guardado localmente como {nombre_archivo_I}")

                        try:
                            contenido_pdf_I = pdf_stream_I.read()
                            informe_codificado_I = base64.b64encode(contenido_pdf_I).decode()
                        except FileNotFoundError:
                            exit()


                        #print(dic_trabajo_I)
                        #Busqueda del ID del equipo en la base de datos maintenance.equipment
                        if tipo_I in intalaciones_interes:
                            serial_I = dic_trabajo_I[f'{i}.2.{equipo} I | N° de serie']
                            try:
                                equipment_I = models.execute_kw(db, uid, password,
                                    'maintenance.equipment', 'search_read',
                                    [[
                                        ['serial_no', '=', serial_I]
                                    ]],
                                    )
                                
                                if equipment_I:

                                    if equipment_I[0]['x_studio_location']:
                                        location_I = equipment_I[0]['x_studio_location'][1]
                                    else:
                                        location_I = False

                                    number_equipment_I = equipment_I[0]['id']

                                    attachment_id = models.execute_kw(db, uid, password,
                                        "ir.attachment", "create", [{
                                            "name": nombre_archivo_I,
                                            #"type": "binary",
                                            "datas": informe_codificado_I,
                                            "res_model": 'maintenance.equipment',
                                            "res_id": number_equipment_I,
                                            "mimetype": "application/pdf",

                                         }])
                                    if location_I == False:

                                        puntos_odoo = models.execute_kw(db, uid, password,
                                            'x_maintenance_location', 'search_read',
                                            [[]],
                                            {'fields': ['id', 'x_name']}, 
                                            )
                                        #Lista de diccionarios

                                        id_punto = None
                                        for p in puntos_odoo:
                                            if p['x_name'] == df_con_datos[f'{i}.1 Punto de monitoreo'][0]:
                                                id_punto = p['id']
                                                break

                                        new_location_I = {
                                            'x_studio_location': id_punto,
                                        # 'effective_date': f"{dic_trabajo_I['Fecha visita ']}",
                                        }

                                        try:
                                            update_location_I = models.execute_kw(
                                                db, uid, password,
                                                'maintenance.equipment',
                                                'write',
                                                [
                                                    [number_equipment_I], 
                                                    new_location_I
                                                ],)

                                            star_location = models.execute_kw(db, uid, password,
                                                'maintenance.equipment', 'message_post', [
                                                    number_equipment_I,
                                                ], {
                                                    'body': f"<p>Ubicación asignada: {punto}</p>",
                                                    'message_type': 'comment',
                                                    'subtype_xmlid': 'mail.mt_note',  # nota interna
                                                    'attachment_ids': [attachment_id]
                                                    })

                                            detalle_op(exito, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, serial_I, id,
                                                    f'Se asocia correctamente el dispositico con el punto de monitoreo {punto}')
    
                                        except Exception as e:

                                            try:
                                                star_location = models.execute_kw(db, uid, password,
                                                    'maintenance.equipment', 'message_post', [
                                                        number_equipment_I,
                                                    ], {
                                                        'body': f"<p>Nueva ubicación: {punto}</p>",
                                                        'message_type': 'comment',
                                                        'subtype_xmlid': 'mail.mt_note',  # nota interna
                                                        'attachment_ids': [attachment_id]
                                                     })
                                            except Exception as e:
                                                print(f'Error al notificar la nueva ubicación del equipo en Odoo: {e}')

                                            detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, serial_I, id, 
                                                        f'{punto} no se encuentra listado en Odoo y Connecteam({type(e)})')

                                    
                                        #Incluir esta validación dentro de MC y MP
                                    elif location_I != df_con_datos[f'{i}.1 Punto de monitoreo'][0]:
                                        detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, serial_I, id, 
                                                    f'El dispositivo ahora se encuentra en {punto}')
        
                                        try:
                                            new_location = models.execute_kw(db, uid, password,
                                                'maintenance.equipment', 'message_post', [
                                                    number_equipment_I,
                                                ], {
                                                    'body': f"<p>La ubicación a cambiado.</p><p>Nueva ubicación: {punto}</p>",
                                                    'message_type': 'comment',
                                                    'subtype_xmlid': 'mail.mt_note',  # nota interna
                                                    'attachment_ids': [attachment_id]
                                                })
                                            
                                        except Exception as e:
                                            print(f'Error al notificar la nueva ubicación del equipo en Odoo: {e}')

                                    else:
                                        try:
                                            last_location = models.execute_kw(db, uid, password,
                                                'maintenance.equipment', 'message_post', [
                                                    number_equipment_I,
                                                ], {
                                                    'body': f"<p>Última ubicación: {punto}</p>",
                                                    'message_type': 'comment',
                                                    'subtype_xmlid': 'mail.mt_note',  # nota interna
                                                    'attachment_ids': [attachment_id]
                                                })
                                            detalle_op(exito, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, serial_I, id,
                                                    'Equipo se mantiene en la misma ubicación')
                                        except Exception as e:
                                            print(f'Error al notificar la ubicación del equipo en Odoo: {e}')

                                else:
                                    try:
                                        pdf_stream_I.seek(0)
                                        sp.upload_file(f'https://graph.microsoft.com/v1.0/drives/b!dx9RXh45RU6gEd39TWLgKItDBbzJweRPoWAkjonKJ4GcIDolNOD0TI7SvyLL7Hda/root:/04. Instalación y Mantenimiento/Trazabilidad de mantenciones y calibraciones/Informes de instalación/{nombre_archivo_I}:/content', pdf_stream_I, "application/pdf" )

                                    except Exception as e:
                                        print(f"Error al subir el informe al Sharepoint: {e}")

                                    detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, serial_I, id,
                                                f'N° de serie no encontrado en Odoo. | Revisar OT {nombre_archivo_I}')
                                    continue
                                            
                            except Exception as e:
                                print(f"Error al buscar equipo en base de Odoo I: {type(e)}")
                                traceback.print_exc()
                                
                                continue
                        else:
                            detalle_op(resumen, ot, tecnico, fecha, proyecto, punto, tipo_I, modelo_I, 'No solicitado', id,
                                    f'Equipo/instrumento no considerado dentro del modulo de mantención')

def check_new_sub(ordered_responses):    
    """
    Procesa una lista de respuestas ordenadas para identificar y registrar nuevas OTs (órdenes de trabajo) en una base de datos SQLite.
    Parámetros:
        ordered_responses (list): Lista de objetos tipo DataFrame o diccionario que contienen información de OTs. 
                                  Se espera que cada elemento tenga una columna o clave '#' con el ID de la OT.
    Funcionamiento:
        1. Verifica si la lista 'ordered_responses' está vacía. Si es así, imprime un mensaje y retorna.
        2. Extrae los IDs de las OTs de la columna '#' del primer elemento de cada DataFrame/diccionario.
        3. Convierte los IDs extraídos a enteros y los almacena en un conjunto para evitar duplicados.
        4. Abre una conexión a la base de datos SQLite 'form_entries.db'.
        5. Prepara una consulta SQL segura usando placeholders para buscar los IDs ya procesados en la tabla 'processed_entries'.
        6. Ejecuta la consulta y obtiene los IDs ya registrados, almacenándolos en un conjunto.
        7. Calcula la diferencia entre los IDs extraídos y los ya procesados para identificar las nuevas OTs.
        8. Si no hay nuevas OTs, imprime un mensaje y retorna.
        9. Filtra los elementos de 'ordered_responses' que corresponden a las nuevas OTs.
        10. Por cada nueva OT, inserta su ID en la tabla 'processed_entries' usando 'INSERT OR IGNORE' para evitar duplicados.
        11. Imprime un mensaje por cada OT guardada en la base de datos.
        12. Retorna la lista de nuevas entradas procesadas.
        13. Si ocurre un error con la base de datos, imprime el error y retorna una lista vacía.
    Retorna:
        list: Lista de nuevas entradas (OTs) que fueron identificadas y registradas en la base de datos.
    """

    if not ordered_responses:
       print("No se encontraron nuevas OTs para procesar.") 
       return
    
    ots = (df['#'][0] for df in ordered_responses)
    ots_id = {int(i) for i in ots}
    
    try:
        with sqlite3.connect('form_entries.db') as connection:
            cursor = connection.cursor()

            # Creamos placeholders (?,?,?) para una consulta segura
            if ots_id:
                placeholders = ','.join(['?'] * len(list(ots_id)))
                query = f"SELECT entry_id FROM processed_entries WHERE entry_id IN ({placeholders})"
                cursor.execute(query, tuple(ots_id))
                # El resultado es una lista de tuplas (e.g., [(101,), (102,)]), las convertimos a un set.
                processed_ids = {row[0] for row in cursor.fetchall()}
            else:
                processed_ids = set()
    

            # Comparar para encontrar solo lo nuevo
            new_ids = ots_id - processed_ids
            if not new_ids:
                print("No hay nuevas OTs para procesar.")
                return

            new_entries = [i for i in ordered_responses if i["#"][0] in new_ids]

            # Registrar en la base de datos el ID de las nuevas OT encontradas
            for entry in new_entries:
                cursor.execute("INSERT OR IGNORE INTO processed_entries (entry_id) VALUES (?)", (int(entry['#'][0]),))
                print(f"ID {entry['#'][0]} guardado en la base de datos.")
            
            return new_entries
    
    except sqlite3.Error as e:
        print(f'Error en la base de datos: {e}')
        traceback.print_exc
        return []

def send_data(df, sheet, table):
    manual = df.values.tolist()
    if manual != []:
        modify_excel_file(manual, sheet, table)

def job(API_key_c):
    try:
        # Obtiene la estructura del formulario y las submissions filtradas, luego las ordena
        ordered_responses_2 = ordenar_respuestas(form_structure(API_key_c), filter_submissions(API_key_c))
    except Exception as e:
        # Si ocurre un error en la conexión a la API, lo muestra
        print(f"Ocurrio un problema con la conexión a la API-Connecteam: {e}")

    try:
        # Busca nuevas OTs que no hayan sido procesadas previamente
        nuevas_entradas = check_new_sub(ordered_responses_2)
        if nuevas_entradas:
            # Inicializa los diccionarios para resumen y éxito de operaciones
            resumen = {
                'OT': [], 'Técnico': [], 'Fecha de revisión': [], 'Proyecto': [],
                'Punto de monitoreo': [], 'Equipo/instrumento': [], 'Modelo': [],
                'N° serie': [], 'Tipo': [], 'Mensaje': []
            }
            exito = {
                'OT': [], 'Técnico': [], 'Fecha de revisión': [], 'Proyecto': [],
                'Punto de monitoreo': [], 'Equipo/instrumento': [], 'Modelo': [],
                'N° serie': [], 'Tipo': [], 'Mensaje': []
            }

            print(f"Se encontraron {len(nuevas_entradas)} nuevas entradas. Procesando...")
            # Procesa las nuevas entradas encontradas
            process_entrys(nuevas_entradas, API_key_c, resumen, exito)

            # Convierte los diccionarios a DataFrames para facilitar el manejo de datos
            df_resumen = pd.DataFrame(resumen)
            df_exito = pd.DataFrame(exito)

            # Une los resultados de resumen y éxito en un DataFrame general
            df_general = pd.DataFrame({k: resumen.get(k, []) + exito.get(k, []) for k in resumen.keys()})

            # Filtra las operaciones que requieren tratamiento manual por tipo de trabajo
            df_manual_m = df_resumen[(df_resumen['Tipo'] == 'MC') | (df_resumen['Tipo'] == 'MP')]
            df_manual_i = df_resumen[df_resumen['Tipo'] == 'I']

            try:
                # Envía los datos filtrados a SharePoint, actualizando los archivos correspondientes
                send_data(df_manual_i, 'Instalaciones', 'resumen_instalación')
                send_data(df_manual_m, 'Mantenciones', 'resumen_mantenciones')

                # Muestra por consola los resúmenes de operaciones manuales y exitosas
                print("\nResumen de operaciones para tratamiento manual:")
                print(tabulate.tabulate(df_resumen, headers='keys', tablefmt='grid'))

                print("\nResumen de operaciones exitosas:")
                print(tabulate.tabulate(df_exito, headers='keys', tablefmt='grid'))

            except Exception as e:
                # Si ocurre un error al actualizar SharePoint, lo muestra
                print(f"Error al actualizar sharepoint: {e}")


    except Exception as e:
        # Si ocurre un error durante la ejecución de la tarea, lo muestra
        print(f"Ocurrió un error durante la ejecución de la tarea: {e}")




#-----------------------------------------------------------------------------------------
#CONEXIÓN API 365
load_dotenv()

sp = Sharepoint()

#-----------------------------------------------------------------------------------------
#CONEXIÓN CON API ODOO
#Datos de conexión
os.environ['SSL_CERT_FILE'] = certifi.where()

url = os.getenv('URL_Odoo')
db = os.getenv('DB_Odoo')
username = os.getenv('USER_Odoo')
password = os.getenv('ODOO_API_KEY')

# #Inicio de sesión
try: 
    common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
    uid = common.authenticate(db, username, password, {})
except Exception as e:
    print(f"Error en el inicio de sesión desde la API: {e}")
    

#Inicialización del endpoint
try:
    models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))
except Exception as e:
    print(f"Error en la inicialización del endpoint: {e}")


#-----------------------------------------------------------------------------------------
#CONEXIÓN CON API CONNECTEAM

API_key_c = os.getenv('CONNECTEAM_API_KEY')

#-----------------------------------------------------------------------------------------
#CONEXIÓN CON API SHAREPOINT
USERNAME = os.getenv('sharepoint_user')
PASSWORD = os.getenv('sharepoint_password')
SHAREPOINT_SITE = os.getenv('sharepoint_url_site')
SHAREPOINT_NAME_SITE = os.getenv('sharepoint_site_name')
SHAREPOINT_DOC = os.getenv('sharepoint_doc_library')

job(API_key_c)





    
    

    







